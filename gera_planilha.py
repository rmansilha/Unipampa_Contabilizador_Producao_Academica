#!/usr/bin/python3
# -*- coding: iso-8859-15 -*-

__author__ = 'Rodrigo Mansilha'
__email__ = 'mansilha@unipampa.edu.br'
__version__ = '{0}.{0}.{1}'
__credits__ = ['Comissão Local de Pesquisa (CLP) do Campus Alegrete da Unipampa - Gestão 2019-2020']

# Bibliotecas gerais

# Sistema
import sys
import os
# parse arguments
import argparse
# Unix style pathname pattern expansion
# import glob
# organizar saídas em formato de log
import logging
from pathlib import Path

# Outras bibliotecas

# ler e escrever arquivos em formato do Excel
from openpyxl import load_workbook, Workbook

# CONFIGURAÇÕES DA PLANILHA DE ENTRADA
MAX_COLUNAS = 100
LINHA_ANOS_PADRAO = 5

# PARAMETROS DE ENTRADA PADRÕES
PLANILHA_PADRAO = "Grupo"
DIRETORIO_ENTRADA_PADRAO = "."
ARQUIVO_SAIDA_PADRAO = "producao_grupos_de_pesquisa.xlsx"
ANO_PADRAO = "2018/2019"

# Todos os campos do relatório de gestão atual
campos = [
	"Artigos completos publicados em periódicos",
	"Livros publicados/organizados ou edições",
	"Capítulos de livros publicados",
	"Trabalhos completos publicados em anais de congressos",
	"Resumos expandidos publicados em anais de congressos",
	"Resumos publicados em anais de congressos",
	"Artigos aceitos para publicação",
	"Apresentações de trabalhos",
	"Demais tipos de produção bibliográfica",
	"Softwares sem registro de patente",
	"Trabalhos técnicos",
	"Produtos artísticos",
	"Demais tipos de produção técnica"
]

# Campos sugeridos
campos_resumo_geral = {
	"Artigos completos publicados em periódicos": [(8, 15), (17, 24)],
	"Livros publicados/organizados ou edições": [(35, 37)],
	"Capítulos de livros publicados": [(38, 39)],
	"Trabalhos completos publicados em anais de congressos": [(26, 28)],
	"Resumos expandidos publicados em anais de congressos": [(30, 30)],
	"Resumos publicados em anais de congressos": [(29, 29)]
}

# "Artigos aceitos para publicação", #
# "Apresentações de trabalhos", #
# "Demais tipos de produção bibliográfica",
# "Softwares sem registro de patente",
# "Trabalhos técnicos",
# "Produtos artísticos",
# "Demais tipos de produção técnica"}

# Campos sugeridos para detalhe
campos_detalhe_grupos = {
	"Artigos publicados em periódicos com Qualis Restrito": [(8, 15), (17, 24)],
	"Artigos publicados em periódicos com Qualis Irrestrito": [(8, 15), (17, 24)],
	"Artigos apresentados em eventos com Qualis Restrito": [(8, 15), (17, 24)],
	"Artigos apresentados em eventos com Qualis Irrestrito": [(8, 15), (17, 24)],
	"Resumos apresentados em eventos": [(8, 15), (17, 24)],
	"Livros publicados ou organizados ou editados": [(35, 37)],
	"Capítulos de livros publicados": [(38, 39)],
	"Patente registrada": [(38, 39)],
	"Projetos externos": [(38, 39)],	    # atualmente em unidades, mas seria melhor em R$
	# "Registro de software":[(??,??)], # sem informação no modelo atual
}


def procura_coluna(planilha_, linha_, valor_procurado_):
	"""
	Retorna o indice da coluna com determinado valor dada uma linha

	:param planilha_: openpyxl.sheet
	:param linha_: inteiro
	:param valor_procurado_: string
	:return: indice da coluna
	"""

	for i in range(1, MAX_COLUNAS):
		valor = planilha_.cell(row=linha_, column=i).value
		# print (planilha_.cell(row=linha_, column=i).column) # debug
		logging.debug(planilha_.cell(row=linha_, column=i).column)
		if valor == valor_procurado_:
			return i

	print("ERRO: na linha %d não foi encontrada coluna com valor '%s'" % (linha_, valor_procurado_))
	sys.exit(-1)


def le_dados_linhas(planilha_, coluna_, linha_inicio_, linha_fim_):
	"""
	Lê dados de uma conjunto de célunas e retorna total
	:param planilha_: openpyxl.sheet
	:param coluna_: inteiro
	:param linha_inicio_: inteiro
	:param linha_fim_: inteiro
	:return: somatório
	"""

	total = 0
	for i in range(linha_inicio_, linha_fim_):
		try:
			valor = planilha_.cell(row=i, column=coluna_).value
			if valor is None:
				valor = 0
			else:
				valor = int(valor)

		except Exception as e:
			print(e)
			valor = planilha_.cell(row=i, column=coluna_).value
			logging.error("Linha:%02d Coluna:%02d Valor:%s " % (i, coluna_, valor))
			sys.exit(-1)

		total += valor
		logging.debug("Linha:%02d Coluna:%02d Valor:%03d Total:%03d" % (i, coluna_, valor, total))

	return total


def le_dados_campos(planilha_, coluna_, campos_, resultado_):
	"""
	Lê dados de um conjunto de campos

	:param planilha_: openpyxl.sheet
	:param coluna_: indice
	:param campos_: array de pares ordenados (linha_início, linha_fim)
	:param resultado_:
	:return: dicionário com valores
	"""

	for campo in campos_.keys():
		sequencias_de_linhas = campos_[campo]
		logging.debug("Campo:%s Sequencias_de_linhas:%s" % (campo, sequencias_de_linhas))  # debug
		total = 0
		for sequencia_de_linhas in sequencias_de_linhas:
			total += le_dados_linhas(planilha_, coluna_, sequencia_de_linhas[0], sequencia_de_linhas[1])

		if campo in resultado_:
			resultado_[campo] += total
		else:
			resultado_[campo] = total

	return resultado_


def exporta_dados_xlsx(planilha_, dados_detalhe_grupos_,):
	"""
	Preenche planilha com dados

	:param planilha_: openpyxl.sheet
	:param dados_detalhe_grupos_: dicionário com valores
	"""

	linha = 0
	for grupo in dados_detalhe_grupos_.keys():
		linha += 2
		logging.info(grupo)
		planilha_.cell(row=linha, column=1).value = grupo

		dados = dados_detalhe_grupos_[grupo]
		logging.debug("Dados: %s" % dados_detalhe_grupos_[grupo])
		largura_coluna = max(len(campo) for campo in dados.keys())  # padding

		for campo in dados.keys():
			logging.info("%s | %03d" % (campo.ljust(largura_coluna), dados[campo]))
			linha += 1
			planilha_.cell(row=linha, column=1).value = campo
			planilha_.cell(row=linha, column=2).value = dados[campo]
		logging.info("")


def exporta_dados_referencia_cruzada_xlsx(planilha_, dados_detalhe_grupos_):
	"""
	Preenche planilha com dados em tabela única (i.e. referência cruzada)

	:param planilha_: openpyxl.sheet
	:param dados_detalhe_grupos_: dicionário com valores
	"""
	coluna = 1
	linha = 1

	# imprime primeira coluna com os campos
	planilha_.cell(row=linha, column=coluna).value = "Grupo"
	grupo0 = list(dados_detalhe_grupos_.keys())[0]
	dados = dados_detalhe_grupos_[grupo0]
	for campo in dados.keys():
		coluna += 1
		planilha_.cell(row=linha, column=coluna).value = campo

	# imprime proximas colunas com os dados dos grupos
	for grupo in dados_detalhe_grupos_.keys():
		coluna = 1
		linha += 1
		planilha_.cell(row=linha, column=coluna).value = grupo
		for campo in dados.keys():
			coluna += 1
			planilha_.cell(row=linha, column=coluna).value = dados[campo]
		logging.info("")


def cria_estrutura_diretorios(planilha_):
	"""

	:param planilha_:
	:return:
	"""
	coluna = 2
	linha = 3

	# imprime primeira coluna com os campos
	planilha_.cell(row=linha, column=coluna).value = "Grupo"
	sigla = planilha_.cell(row=linha, column=coluna).value
	while sigla:
		try:
			diretorio = "./GruposDePesquisa/%s" % sigla
			logging.debug("\tDiretório: %s" % diretorio)
			os.makedirs(diretorio, exist_ok=True)
			linha += 1
			sigla = planilha_.cell(row=linha, column=coluna).value

		except Exception as e:
			print(e)


def main():
	"""
	Programa principal
	"""

	# Configura argumentos
	parser = argparse.ArgumentParser(description='Processa produção científica de grupos de Pesquisa da UNIPAMPA.')
	parser.add_argument("--dir", "-d", help="diretório com arquivos de entrada.", default=DIRETORIO_ENTRADA_PADRAO)
	parser.add_argument("--saida", "-s", help="arquivo de saída.", default=ARQUIVO_SAIDA_PADRAO)
	parser.add_argument("--ano", "-a", help="ano de processamento.", default=ANO_PADRAO)
	parser.add_argument("--entrada", "-e", help="arquivo de entrada com grupos de pesquisa. Gera estrutura de diretórios.", default="")
	help_log = "nível de log (INFO=%d DEBUG=%d)" % (logging.INFO, logging.DEBUG)
	parser.add_argument("--log", "-l", help=help_log, default=logging.INFO, type=int)
	parser.print_help()

	# lê argumentos da linha de comando
	args = parser.parse_args()

	# configura log
	# logging.basicConfig(level=args.log, format='%(asctime)s - %(message)s')
	logging.basicConfig(level=args.log, format='%(message)s')

	# mostra parâmetros de entrada
	logging.info("")
	logging.info("PARÂMETROS DE ENTRADA")
	logging.info("---------------------")
	logging.info("\tentrada : %s" % args.entrada)

	if args.entrada:
		logging.info("")
		logging.info("CRIANDO DIRETÓRIOS")
		logging.info("------------------")
		workbook_entrada = load_workbook(filename=args.entrada)
		planilha = workbook_entrada.active
		cria_estrutura_diretorios(planilha)

	else:
		logging.info("\t log     : %d" % args.log)
		logging.info("\t saída   : %s" % args.saida)
		logging.info("\t ano     : %s" % args.ano)
		logging.info("\t dir     : %s" % args.dir)

		# mostra parâmetros calculados
		logging.info("")
		logging.info("PARÂMETROS CALCULADOS")
		logging.info("---------------------")
		# arquivos = [arquivo for arquivo in glob.glob("%s/[!~]*.xlsx"%args.dir)]
		arquivos = [arquivo for arquivo in Path(args.dir).rglob('[!~]*.xlsx')]

		if Path(args.saida) in arquivos:
			arquivos.remove(Path(args.saida))

		logging.info("\t arquivos: %s" % arquivos)

		# inicializa variáveis
		dados_resumo_geral = {}
		dados_resumo_ano = {}
		dados_resumo_ano["Ano: %s" % args.ano] = dados_resumo_geral
		dados_detalhe_grupos = {}
		conta_arquivo = 1

		# processa dados
		logging.info("")
		logging.info("LEITURA DE DADOS")
		logging.info("----------------")
		try:
			for nome_arquivo in arquivos:
				base_nome_arquivo = os.path.basename(nome_arquivo)
				nome_grupo = os.path.splitext(base_nome_arquivo)[0]
				logging.info("\t(%d/%d) Arquivo: %s Grupo: %s" % (conta_arquivo, len(arquivos), nome_arquivo, nome_grupo))
				workbook = load_workbook(filename=nome_arquivo)
				planilha = workbook[PLANILHA_PADRAO]
				coluna = procura_coluna(planilha, LINHA_ANOS_PADRAO, args.ano)

				# GERAL
				dados_resumo_geral = le_dados_campos(planilha, coluna, campos_resumo_geral, dados_resumo_geral)

				# GRUPO
				dados_grupo = le_dados_campos(planilha, coluna, campos_detalhe_grupos, {})
				dados_detalhe_grupos["%s" % nome_grupo] = dados_grupo

		except Exception as e:
			print(e)
			sys.exit(-1)

		logging.info("")
		logging.info("RESULTADOS")
		logging.info("----------")
		logging.debug("GERAL")
		logging.debug("\t", dados_resumo_geral)
		logging.debug("")
		logging.debug("GRUPOS")
		logging.debug("\t", dados_detalhe_grupos)

		# Grava resultados
		workbook_saida = Workbook()
		workbook_saida.remove(workbook_saida.active)
		planilha_saida = workbook_saida.create_sheet("resumo")
		exporta_dados_xlsx(planilha_saida, dados_resumo_ano)
		logging.debug("")

		planilha_saida = workbook_saida.create_sheet("detalhes")
		exporta_dados_xlsx(planilha_saida, dados_detalhe_grupos)

		planilha_saida = workbook_saida.create_sheet("detalhes_unica_tabela")
		exporta_dados_referencia_cruzada_xlsx(planilha_saida, dados_detalhe_grupos)
		workbook_saida.save(args.saida)

	logging.info("\nPronto.\n")


if __name__ == "__main__":
	main()
